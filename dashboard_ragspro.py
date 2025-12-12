#!/usr/bin/env python3
"""
RAGSPRO DASHBOARD - Complete Dark Theme Dashboard with All Features
Port: 5002
"""

from flask import Flask, render_template, jsonify, request, send_file
import sys
import os
import json
from datetime import datetime
import threading
import time
import logging
from functools import wraps

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Configure Flask
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')

# Rate Limiting
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Global state with thread safety
generation_status = {
    'running': False,
    'progress': 0,
    'current_query': '',
    'leads_found': 0,
    'message': 'Ready to generate premium leads',
    'last_run': None,
    'stop_signal': False,
    'latest_leads': []
}

# Thread locks for safety
status_lock = threading.Lock()
file_lock = threading.Lock()
generation_lock = threading.Lock()


def load_premium_leads():
    """Load premium leads from JSON file with thread safety."""
    json_path = "data/premium_leads.json"
    
    # Ensure data directory exists
    os.makedirs("data", exist_ok=True)
    
    with file_lock:  # Thread-safe file access
        if not os.path.exists(json_path):
            logger.warning(f"Premium leads file not found at {json_path}, creating empty file")
            # Create empty leads file
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump([], f)
            return []
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if not content:
                    logger.warning("Premium leads file is empty")
                    return []
                leads = json.loads(content)
            logger.info(f"Loaded {len(leads)} leads from {json_path}")
            return leads
        except Exception as e:
            logger.error(f"Error loading leads: {e}")
            return []


def save_premium_leads(leads, append=False):
    """Save premium leads to JSON file with thread safety and atomic writes."""
    os.makedirs("data", exist_ok=True)
    os.makedirs("data/history", exist_ok=True)
    
    json_path = "data/premium_leads.json"
    temp_path = json_path + ".tmp"
    
    with file_lock:  # Thread-safe file access
        try:
            if append and os.path.exists(json_path):
                with open(json_path, 'r', encoding='utf-8') as f:
                    existing_leads = json.load(f)
                
                # Remove duplicates
                seen = set()
                combined = []
                for lead in existing_leads + leads:
                    key = (lead.get('title', ''), lead.get('address', ''))
                    if key not in seen:
                        seen.add(key)
                        combined.append(lead)
                leads = combined
            
            # Atomic write: write to temp file first, then rename
            with open(temp_path, 'w', encoding='utf-8') as f:
                json.dump(leads, f, indent=2, ensure_ascii=False)
            
            # Atomic rename (prevents corruption)
            os.replace(temp_path, json_path)
            
            # Save to history
            today = datetime.now().strftime('%Y-%m-%d')
            history_path = f"data/history/leads_{today}.json"
            history_data = {
                'date': today,
                'timestamp': datetime.now().isoformat(),
                'total_leads': len(leads),
                'leads': leads
            }
            with open(history_path, 'w', encoding='utf-8') as f:
                json.dump(history_data, f, indent=2, ensure_ascii=False)
            
            logger.info(f"Saved {len(leads)} leads")
            return True
        except Exception as e:
            logger.error(f"Error saving leads: {e}")
            # Clean up temp file if it exists
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
            return False


def generate_ai_content(lead):
    """Generate AI content for a lead."""
    try:
        from src.ai_gemini import create_ai_assistant
        from src.config import load_config
        
        config = load_config()
        if not config.get('GEMINI_API_KEY'):
            return generate_fallback_content(lead)
        
        ai = create_ai_assistant(config['GEMINI_API_KEY'])
        
        email = ai.generate_cold_email(
            lead['title'],
            lead.get('type', 'business'),
            lead.get('address', ''),
            lead.get('rating', 0),
            lead.get('reviews', 0)
        )
        
        whatsapp = ai.generate_whatsapp_message(
            lead['title'],
            lead.get('type', 'business')
        )
        
        return {'email': email, 'whatsapp': whatsapp}
    except Exception as e:
        logger.error(f"AI generation error: {e}")
        return generate_fallback_content(lead)


def generate_fallback_content(lead):
    """Generate fallback content without AI."""
    business_name = lead['title']
    business_type = lead.get('type', 'business')
    
    email = f"""Subject: Grow {business_name} Online - Free Consultation

Hi {business_name} Team,

I came across your {business_type} and was impressed!

At RagsPro.com, we help businesses like yours:
✅ Get 3-5x more customers through digital marketing
✅ Build modern, high-converting websites
✅ Rank #1 on Google with proven SEO

Would you be interested in a FREE consultation?

Best regards,
Raghav Sharma
Founder, RagsPro.com
📞 +91-8700048490
🌐 www.ragspro.com"""

    whatsapp = f"""Hi {business_name}! 👋

Noticed your {business_type} - impressive! 🌟

We help businesses get 3-5x more customers through:
✅ Modern websites
✅ Google Ads & SEO
✅ Social media

FREE consultation available! Interested?

- Raghav, RagsPro.com
📞 +91-8700048490"""

    return {'email': email, 'whatsapp': whatsapp}


def run_premium_generation(target_countries, target_cities, business_types, num_leads, quality_threshold):
    """Run premium lead generation with thread safety and user-specified filters."""
    global generation_status
    
    try:
        with status_lock:
            generation_status['running'] = True
            generation_status['stop_signal'] = False
            generation_status['progress'] = 5
            generation_status['message'] = '🚀 Initializing...'
        
        from src.scraper import search_places
        from src.lead_quality_filter import filter_serious_clients_only
        from src.queries import CITIES, CATEGORIES
        from src.filters import remove_duplicates
        from src.config import load_config
        
        config = load_config()
        api_key = config.get('SERPAPI_KEY')
        
        if not api_key:
            generation_status['running'] = False
            generation_status['message'] = '❌ API Key not configured'
            return
        
        with status_lock:
            generation_status['progress'] = 15
            generation_status['message'] = 'Preparing queries...'
        
        # Filter cities based on user selection
        if target_cities and len(target_cities) > 0:
            # User selected specific cities
            filtered_cities = target_cities
            logger.info(f"🎯 Using user-selected cities: {filtered_cities}")
        elif target_countries and len(target_countries) > 0:
            # User selected countries, filter cities by country
            filtered_cities = [city for city in CITIES 
                              if any(country in city for country in target_countries)]
            logger.info(f"🌍 Filtered {len(filtered_cities)} cities for countries: {target_countries}")
        else:
            # No filter, use all cities (limit to 10 for speed)
            filtered_cities = CITIES[:10]
            logger.info(f"🌎 Using all cities (limited to 10)")
        
        # Filter categories based on user selection
        if business_types and len(business_types) > 0:
            # User selected specific business types
            filtered_categories = business_types
            logger.info(f"💼 Using user-selected categories: {filtered_categories}")
        else:
            # No filter, use all categories (limit to 10 for speed)
            filtered_categories = CATEGORIES[:10]
            logger.info(f"🎯 Using all categories (limited to 10)")
        
        # Generate queries from filtered cities and categories
        all_queries = []
        for city in filtered_cities[:20]:  # Max 20 cities
            for category in filtered_categories[:20]:  # Max 20 categories
                query = f"{category} in {city}"
                all_queries.append(query)
        
        logger.info(f"📋 Generated {len(all_queries)} search queries")
        
        with status_lock:
            generation_status['progress'] = 20
            generation_status['message'] = f'Searching {len(all_queries)} locations...'
        
        # Scrape leads
        premium_leads = []
        
        for i, query in enumerate(all_queries):
            with status_lock:
                if generation_status['stop_signal'] or len(premium_leads) >= num_leads:
                    break
                
                progress = 20 + int((i / len(all_queries)) * 60)
                generation_status['progress'] = progress
                generation_status['current_query'] = query
                generation_status['message'] = f'Searching... ({i+1}/{len(all_queries)})'
            
            try:
                results = search_places(query, api_key)
                if not results:
                    continue
                
                quality_leads = filter_serious_clients_only(results)
                premium = [lead for lead in quality_leads 
                          if lead.get('quality_score', 0) >= quality_threshold]
                
                if premium:
                    premium_leads.extend(premium)
                    with status_lock:
                        generation_status['leads_found'] = len(premium_leads)
                        # Limit latest_leads to prevent memory leak
                        generation_status['latest_leads'] = (generation_status['latest_leads'] + premium)[-100:]
                    save_premium_leads(premium, append=True)
                    
            except Exception as e:
                logger.error(f"Error scraping {query}: {e}")
                continue
        
        with status_lock:
            generation_status['progress'] = 90
            generation_status['message'] = 'Removing duplicates...'
        
        unique_leads = remove_duplicates(premium_leads)
        
        # Add timestamp to new leads so we can identify them
        timestamp = datetime.now().isoformat()
        for lead in unique_leads:
            lead['generated_at'] = timestamp
            lead['is_new'] = True
        
        with status_lock:
            generation_status['progress'] = 100
            generation_status['leads_found'] = len(unique_leads)
            generation_status['message'] = f'✅ Complete! Generated {len(unique_leads)} premium leads'
            generation_status['last_run'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        logger.info(f"✅ Generation complete: {len(unique_leads)} leads")
        
    except Exception as e:
        logger.error(f"❌ Generation error: {e}", exc_info=True)
        with status_lock:
            generation_status['message'] = f'❌ Error: {str(e)}'
            generation_status['progress'] = 0
    finally:
        with status_lock:
            generation_status['running'] = False


@app.route('/')
def index():
    """Main RAGSPRO dashboard page."""
    return render_template('ragspro_dashboard.html')


@app.route('/api/debug/files')
def debug_files():
    """Debug endpoint to check file system."""
    try:
        import os
        data_dir = "data"
        files_info = {
            'data_dir_exists': os.path.exists(data_dir),
            'data_dir_contents': os.listdir(data_dir) if os.path.exists(data_dir) else [],
            'premium_leads_exists': os.path.exists('data/premium_leads.json'),
            'premium_leads_size': os.path.getsize('data/premium_leads.json') if os.path.exists('data/premium_leads.json') else 0,
            'cwd': os.getcwd(),
            'app_files': os.listdir('.')[:20]  # First 20 files in app directory
        }
        return jsonify({'success': True, 'files': files_info})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/leads')
def get_leads():
    """Get all premium leads (AI content generated on-demand), sorted with newest first."""
    try:
        leads = load_premium_leads()
        
        # Sort leads: newest first (by generated_at timestamp)
        # This ensures new leads appear at the top
        def get_timestamp(lead):
            return lead.get('generated_at', lead.get('created_at', '2000-01-01'))
        
        leads_sorted = sorted(leads, key=get_timestamp, reverse=True)
        
        # Don't generate AI content here - too slow for 500+ leads
        # Frontend will request it per-lead as needed
        
        return jsonify({
            'success': True,
            'leads': leads_sorted,
            'total': len(leads_sorted)
        })
    except Exception as e:
        logger.error(f"Error getting leads: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/stats')
def get_stats():
    """Get dashboard statistics."""
    try:
        leads = load_premium_leads()
        
        if not leads:
            return jsonify({
                'success': True,
                'stats': {
                    'total_leads': 0,
                    'avg_quality': 0,
                    'avg_rating': 0,
                    'countries': {},
                    'categories': {},
                    'last_run': generation_status.get('last_run', 'Never')
                }
            })
        
        total_leads = len(leads)
        avg_quality = sum(l.get('quality_score', 0) for l in leads) / total_leads
        avg_rating = sum(l.get('rating', 0) for l in leads) / total_leads
        
        # Count by country
        countries = {}
        for lead in leads:
            address = lead.get('address', '')
            country = 'Unknown'
            for c in ['USA', 'UK', 'UAE', 'Canada', 'Australia', 'India']:
                if c in address:
                    country = c
                    break
            countries[country] = countries.get(country, 0) + 1
        
        # Count by category
        categories = {}
        for lead in leads:
            cat = lead.get('type', 'Unknown')
            categories[cat] = categories.get(cat, 0) + 1
        
        return jsonify({
            'success': True,
            'stats': {
                'total_leads': total_leads,
                'avg_quality': round(avg_quality, 1),
                'avg_rating': round(avg_rating, 2),
                'countries': dict(sorted(countries.items(), key=lambda x: x[1], reverse=True)[:5]),
                'categories': dict(sorted(categories.items(), key=lambda x: x[1], reverse=True)[:5]),
                'last_run': generation_status.get('last_run', 'Never')
            }
        })
    except Exception as e:
        logger.error(f"Error getting stats: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/generate', methods=['POST'])
@limiter.limit("5 per hour")  # Rate limit: 5 generations per hour
def generate_leads():
    """Start premium lead generation."""
    global generation_status
    
    with generation_lock:  # Thread-safe check
        if generation_status['running']:
            return jsonify({'success': False, 'message': 'Generation already running'})
    
    try:
        data = request.json or {}
        
        # Support both 'markets' (frontend) and 'countries' (legacy)
        target_countries = data.get('markets', data.get('countries', []))
        target_cities = data.get('cities', [])
        business_types = data.get('business_types', [])
        num_leads = int(data.get('num_leads', 50))
        quality_threshold = float(data.get('quality_threshold', 70))
        clear_old = data.get('clear_old', False)
        
        # Clear old leads if requested
        if clear_old:
            logger.info("🗑️ Clearing old leads before generation...")
            save_premium_leads([])
        
        thread = threading.Thread(
            target=run_premium_generation,
            args=(target_countries, target_cities, business_types, num_leads, quality_threshold)
        )
        thread.daemon = True
        thread.start()
        
        return jsonify({'success': True, 'message': 'Premium lead generation started'})
    except Exception as e:
        logger.error(f"Error starting generation: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/status')
def get_status():
    """Get generation status with thread safety."""
    with status_lock:  # Thread-safe read
        return jsonify({
            'success': True,
            'is_running': generation_status['running'],
            'progress': generation_status['progress'],
            'current_lead': generation_status.get('current_query', ''),
            'leads_found': generation_status.get('leads_found', 0),
            'message': generation_status.get('message', ''),
            'status': generation_status.copy()  # Return copy to prevent external modification
        })


@app.route('/api/stop', methods=['POST'])
def stop_generation():
    """Stop generation with thread safety."""
    global generation_status
    
    with status_lock:  # Thread-safe write
        if not generation_status['running']:
            return jsonify({'success': False, 'message': 'Generation is not running'})
        
        generation_status['stop_signal'] = True
        generation_status['message'] = '🛑 Stopping...'
    
    return jsonify({'success': True, 'message': 'Stop signal sent'})


@app.route('/api/leads/clear', methods=['POST'])
@limiter.limit("10 per hour")
def clear_all_leads():
    """Clear all leads from database."""
    try:
        # Backup current leads before clearing
        leads = load_premium_leads()
        if leads:
            backup_path = f"data/backups/leads_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            os.makedirs("data/backups", exist_ok=True)
            with open(backup_path, 'w', encoding='utf-8') as f:
                json.dump(leads, f, indent=2, ensure_ascii=False)
            logger.info(f"Backed up {len(leads)} leads to {backup_path}")
        
        # Clear the database
        save_premium_leads([])
        
        return jsonify({
            'success': True,
            'message': f'✅ Cleared {len(leads)} leads (backup saved)',
            'backup_path': backup_path if leads else None
        })
    except Exception as e:
        logger.error(f"Error clearing leads: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/search')
def search_leads():
    """Search leads."""
    try:
        query = request.args.get('q', '').lower()
        leads = load_premium_leads()
        
        if query:
            filtered = []
            for lead in leads:
                if (query in lead.get('title', '').lower() or
                    query in lead.get('type', '').lower() or
                    query in lead.get('address', '').lower()):
                    filtered.append(lead)
            leads = filtered
        
        # Don't generate AI content here - frontend will request per-lead
        
        return jsonify({'success': True, 'leads': leads, 'total': len(leads)})
    except Exception as e:
        logger.error(f"Error searching: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/lead/<int:lead_id>/ai-content')
@limiter.limit("30 per minute")  # Rate limit: 30 AI requests per minute
def get_lead_ai_content(lead_id):
    """Generate AI content for a specific lead with caching."""
    try:
        leads = load_premium_leads()
        
        if lead_id >= len(leads):
            return jsonify({'success': False, 'error': 'Lead not found'})
        
        lead = leads[lead_id]
        
        # Check if AI content already exists (cached)
        if lead.get('ai_content'):
            logger.info(f"Using cached AI content for lead {lead_id}")
            return jsonify({'success': True, 'ai_content': lead['ai_content'], 'lead': lead, 'cached': True})
        
        # Generate AI content
        logger.info(f"Generating new AI content for lead {lead_id}")
        ai_content = generate_ai_content(lead)
        
        # Save it back with timestamp
        lead['ai_content'] = ai_content
        lead['ai_generated_at'] = datetime.now().isoformat()
        save_premium_leads(leads)
        
        return jsonify({'success': True, 'ai_content': ai_content, 'lead': lead, 'cached': False})
    except Exception as e:
        logger.error(f"Error generating AI content: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/export/csv', methods=['GET', 'POST'])
def export_csv():
    """Export leads to CSV."""
    try:
        import csv
        import io
        from flask import make_response
        
        # Get selected lead IDs from POST request
        if request.method == 'POST':
            data = request.json
            lead_ids = data.get('lead_ids', [])
            all_leads = load_premium_leads()
            leads = [all_leads[i] for i in lead_ids if i < len(all_leads)]
        else:
            leads = load_premium_leads()
        
        if not leads:
            return "No leads to export", 404
        
        si = io.StringIO()
        cw = csv.writer(si)
        
        cw.writerow(['Business Name', 'Type', 'Location', 'Rating', 'Reviews', 'Phone', 'Website', 'Quality Score', 'Status'])
        
        for lead in leads:
            cw.writerow([
                lead.get('title', ''),
                lead.get('type', ''),
                lead.get('address', ''),
                lead.get('rating', ''),
                lead.get('reviews', ''),
                lead.get('phone', ''),
                lead.get('website', ''),
                lead.get('quality_score', ''),
                lead.get('status', 'New')
            ])
        
        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename=ragspro_leads_{len(leads)}.csv"
        output.headers["Content-type"] = "text/csv"
        return output
    except Exception as e:
        logger.error(f"Export error: {e}")
        return str(e), 500


@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """Export selected leads to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        from flask import make_response
        
        data = request.json
        lead_ids = data.get('lead_ids', [])
        
        all_leads = load_premium_leads()
        leads = [all_leads[i] for i in lead_ids if i < len(all_leads)]
        
        if not leads:
            return jsonify({'success': False, 'error': 'No leads selected'}), 400
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Premium Leads"
        
        # Header style
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Headers
        headers = ['Business Name', 'Type', 'Location', 'Rating', 'Reviews', 'Phone', 'Website', 'Quality Score', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        for row, lead in enumerate(leads, 2):
            ws.cell(row=row, column=1, value=lead.get('title', ''))
            ws.cell(row=row, column=2, value=lead.get('type', ''))
            ws.cell(row=row, column=3, value=lead.get('address', ''))
            ws.cell(row=row, column=4, value=lead.get('rating', ''))
            ws.cell(row=row, column=5, value=lead.get('reviews', ''))
            ws.cell(row=row, column=6, value=lead.get('phone', ''))
            ws.cell(row=row, column=7, value=lead.get('website', ''))
            ws.cell(row=row, column=8, value=lead.get('quality_score', ''))
            ws.cell(row=row, column=9, value=lead.get('status', 'New'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        output = make_response(excel_file.read())
        output.headers["Content-Disposition"] = f"attachment; filename=ragspro_leads_{len(leads)}.xlsx"
        output.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        logger.info(f"✅ Exported {len(leads)} leads to Excel")
        return output
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    """Export selected leads to PDF."""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        import io
        from flask import make_response
        
        data = request.json
        lead_ids = data.get('lead_ids', [])
        
        all_leads = load_premium_leads()
        leads = [all_leads[i] for i in lead_ids if i < len(all_leads)]
        
        if not leads:
            return jsonify({'success': False, 'error': 'No leads selected'}), 400
        
        # Create PDF
        pdf_file = io.BytesIO()
        doc = SimpleDocTemplate(pdf_file, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#7C3AED'),
            spaceAfter=30,
            alignment=1
        )
        
        # Title
        title = Paragraph("RAGSPRO - Premium Leads Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary
        summary_text = f"<b>Total Leads:</b> {len(leads)}<br/><b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary = Paragraph(summary_text, styles['Normal'])
        elements.append(summary)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        table_data = [['Business', 'Type', 'Location', 'Rating', 'Quality']]
        
        for lead in leads:
            table_data.append([
                lead.get('title', '')[:30],
                lead.get('type', '')[:20],
                lead.get('address', '')[:30],
                f"{lead.get('rating', 0)}⭐",
                f"{lead.get('quality_score', 0)}/100"
            ])
        
        # Create table
        table = Table(table_data, colWidths=[2*inch, 1.5*inch, 2*inch, 0.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C3AED')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        pdf_file.seek(0)
        
        output = make_response(pdf_file.read())
        output.headers["Content-Disposition"] = f"attachment; filename=ragspro_leads_{len(leads)}.pdf"
        output.headers["Content-type"] = "application/pdf"
        
        logger.info(f"✅ Exported {len(leads)} leads to PDF")
        return output
        
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/leads/hot')
def get_hot_leads():
    """Get hot leads (quality score > 85)."""
    try:
        leads = load_premium_leads()
        hot_leads = [lead for lead in leads if lead.get('quality_score', 0) > 85]
        return jsonify({'success': True, 'leads': hot_leads, 'total': len(hot_leads)})
    except Exception as e:
        logger.error(f"Error getting hot leads: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/leads/today')
def get_todays_leads():
    """Get today's leads."""
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        history_path = f"data/history/leads_{today}.json"
        
        if os.path.exists(history_path):
            with open(history_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return jsonify({'success': True, 'leads': data.get('leads', [])})
        else:
            return jsonify({'success': True, 'leads': []})
    except Exception as e:
        logger.error(f"Error getting today's leads: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/history')
def get_history():
    """Get generation history."""
    try:
        history_dir = "data/history"
        if not os.path.exists(history_dir):
            return jsonify({'success': True, 'history': []})
        
        history_files = sorted([f for f in os.listdir(history_dir) if f.startswith('leads_') and f.endswith('.json')], reverse=True)
        
        history = []
        for filename in history_files[:30]:  # Last 30 days
            try:
                with open(os.path.join(history_dir, filename), 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    history.append({
                        'date': data.get('date', filename.replace('leads_', '').replace('.json', '')),
                        'total_leads': data.get('total_leads', len(data.get('leads', []))),
                        'timestamp': data.get('timestamp', '')
                    })
            except:
                continue
        
        return jsonify({'success': True, 'history': history})
    except Exception as e:
        logger.error(f"Error getting history: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/history/all')
def get_all_history():
    """Get all historical leads."""
    try:
        history_dir = "data/history"
        if not os.path.exists(history_dir):
            return jsonify({'success': True, 'leads': [], 'total': 0})
        
        all_leads = []
        history_files = sorted([f for f in os.listdir(history_dir) if f.startswith('leads_') and f.endswith('.json')], reverse=True)
        
        for filename in history_files:
            try:
                with open(os.path.join(history_dir, filename), 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    leads = data.get('leads', [])
                    all_leads.extend(leads)
            except:
                continue
        
        return jsonify({'success': True, 'leads': all_leads, 'total': len(all_leads)})
    except Exception as e:
        logger.error(f"Error getting all history: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/send-whatsapp', methods=['POST'])
def send_whatsapp():
    """Send WhatsApp message."""
    try:
        data = request.json
        lead_id = data.get('lead_id')
        phone = data.get('phone', '')
        message = data.get('message', '')
        business_name = data.get('business_name', '')
        
        if not phone or not message:
            return jsonify({'success': False, 'error': 'Phone and message required'})
        
        # Clean phone number
        phone_clean = ''.join(filter(str.isdigit, phone))
        if not phone_clean.startswith('91'):
            phone_clean = '91' + phone_clean
        
        # Create WhatsApp URL
        import urllib.parse
        message_encoded = urllib.parse.quote(message)
        whatsapp_url = f"https://wa.me/{phone_clean}?text={message_encoded}"
        
        # Update lead status
        if lead_id is not None:
            leads = load_premium_leads()
            if lead_id < len(leads):
                leads[lead_id]['status'] = 'WhatsApp Sent'
                leads[lead_id]['last_contacted'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                save_premium_leads(leads)
        
        logger.info(f"✅ WhatsApp sent to {business_name}")
        
        return jsonify({
            'success': True,
            'url': whatsapp_url,
            'message': f'✅ WhatsApp ready for {business_name}'
        })
    except Exception as e:
        logger.error(f"WhatsApp error: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/send-email', methods=['POST'])
def send_email():
    """Send email."""
    try:
        data = request.json
        lead_id = data.get('lead_id')
        email = data.get('email', '')
        subject = data.get('subject', 'Partnership Opportunity - RagsPro.com')
        body = data.get('body', '')
        business_name = data.get('business_name', '')
        
        if not body:
            return jsonify({'success': False, 'error': 'Email body required'})
        
        # Create mailto URL
        import urllib.parse
        subject_encoded = urllib.parse.quote(subject)
        body_encoded = urllib.parse.quote(body)
        mailto_url = f"mailto:{email}?subject={subject_encoded}&body={body_encoded}"
        
        # Update lead status
        if lead_id is not None:
            leads = load_premium_leads()
            if lead_id < len(leads):
                leads[lead_id]['status'] = 'Email Sent'
                leads[lead_id]['last_contacted'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                save_premium_leads(leads)
        
        logger.info(f"✅ Email sent to {business_name}")
        
        return jsonify({
            'success': True,
            'url': mailto_url,
            'message': f'✅ Email ready for {business_name}'
        })
    except Exception as e:
        logger.error(f"Email error: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/bulk/email', methods=['POST'])
def bulk_send_email():
    """Bulk send emails to selected leads."""
    try:
        data = request.json
        lead_ids = data.get('lead_ids', [])
        
        all_leads = load_premium_leads()
        selected_leads = [all_leads[i] for i in lead_ids if i < len(all_leads)]
        
        if not selected_leads:
            return jsonify({'success': False, 'error': 'No leads selected'})
        
        # Generate AI content for each lead
        results = []
        for lead in selected_leads:
            ai_content = generate_ai_content(lead)
            results.append({
                'business_name': lead['title'],
                'email': ai_content['email'],
                'phone': lead.get('phone', '')
            })
        
        return jsonify({
            'success': True,
            'total': len(results),
            'emails': results,
            'message': f'✅ Generated emails for {len(results)} leads'
        })
        
    except Exception as e:
        logger.error(f"Bulk email error: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/bulk/whatsapp', methods=['POST'])
def bulk_send_whatsapp():
    """Bulk send WhatsApp to selected leads."""
    try:
        data = request.json
        lead_ids = data.get('lead_ids', [])
        
        all_leads = load_premium_leads()
        selected_leads = [all_leads[i] for i in lead_ids if i < len(all_leads)]
        
        if not selected_leads:
            return jsonify({'success': False, 'error': 'No leads selected'})
        
        # Generate WhatsApp URLs
        results = []
        for lead in selected_leads:
            phone = lead.get('phone', '')
            if phone:
                ai_content = generate_ai_content(lead)
                phone_clean = ''.join(filter(str.isdigit, phone))
                if not phone_clean.startswith('91'):
                    phone_clean = '91' + phone_clean
                
                import urllib.parse
                message_encoded = urllib.parse.quote(ai_content['whatsapp'])
                whatsapp_url = f"https://wa.me/{phone_clean}?text={message_encoded}"
                
                results.append({
                    'business_name': lead['title'],
                    'phone': phone,
                    'url': whatsapp_url
                })
        
        return jsonify({
            'success': True,
            'total': len(results),
            'whatsapp_urls': results,
            'message': f'✅ Generated WhatsApp for {len(results)} leads'
        })
        
    except Exception as e:
        logger.error(f"Bulk WhatsApp error: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/bulk/linkedin', methods=['POST'])
def bulk_linkedin_search():
    """Bulk LinkedIn search for selected leads."""
    try:
        data = request.json
        lead_ids = data.get('lead_ids', [])
        
        all_leads = load_premium_leads()
        selected_leads = [all_leads[i] for i in lead_ids if i < len(all_leads)]
        
        if not selected_leads:
            return jsonify({'success': False, 'error': 'No leads selected'})
        
        # Generate LinkedIn search URLs
        results = []
        for lead in selected_leads:
            business_name = lead['title']
            location = lead.get('address', '').split(',')[0]
            
            import urllib.parse
            search_query = f"{business_name} {location}"
            linkedin_url = f"https://www.linkedin.com/search/results/companies/?keywords={urllib.parse.quote(search_query)}"
            
            results.append({
                'business_name': business_name,
                'search_url': linkedin_url
            })
        
        return jsonify({
            'success': True,
            'total': len(results),
            'linkedin_urls': results,
            'message': f'✅ Generated LinkedIn searches for {len(results)} leads'
        })
        
    except Exception as e:
        logger.error(f"Bulk LinkedIn error: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/lead/analyze', methods=['POST'])
@limiter.limit("20 per minute")
def analyze_lead():
    """Analyze a lead with AI to identify problems and solutions."""
    try:
        data = request.json
        lead = data.get('lead')
        
        if not lead:
            return jsonify({'success': False, 'error': 'No lead provided'})
        
        # Use AI to analyze the lead
        from src.ai_gemini import create_ai_assistant
        from src.config import load_config
        
        config = load_config()
        if not config.get('GEMINI_API_KEY'):
            return jsonify({
                'success': False,
                'error': 'AI not configured'
            })
        
        ai = create_ai_assistant(config['GEMINI_API_KEY'])
        
        # Generate analysis
        business_name = lead.get('title', '')
        business_type = lead.get('type', '')
        rating = lead.get('rating', 0)
        reviews = lead.get('reviews', 0)
        address = lead.get('address', '')
        
        analysis_text = ai.analyze_business(business_name, business_type, rating, reviews, address)
        
        # Generate quick pitch
        pitch = f"Hi {business_name}! Noticed your {rating}★ rating - impressive! We help {business_type} businesses get 3-5x more customers through modern tech. Interested in a FREE consultation?"
        
        return jsonify({
            'success': True,
            'analysis': analysis_text,
            'quick_pitch': pitch
        })
        
    except Exception as e:
        logger.error(f"Lead analysis error: {e}")
        return jsonify({'success': False, 'error': str(e)})


if __name__ == '__main__':
    print("""
    ╔══════════════════════════════════════════════════════════╗
    ║           RAGSPRO DASHBOARD - DARK THEME                 ║
    ║                                                          ║
    ║  🎯 Complete Lead Management System                      ║
    ║  💰 AI-Powered Content Generation                        ║
    ║  🚀 Real-time Lead Generation                            ║
    ╚══════════════════════════════════════════════════════════╝
    
    🚀 Dashboard running at: http://localhost:5002
    📊 Open your browser and start generating premium leads!
    """)
    
    app.run(debug=True, host='0.0.0.0', port=5002)
